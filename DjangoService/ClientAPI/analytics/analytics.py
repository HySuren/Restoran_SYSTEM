from django.http import JsonResponse
from django.utils.timezone import localtime
from openpyxl.styles import PatternFill
from ClientAPI.models import Order
import openpyxl
from pathlib import Path

import os

# Build paths inside the project like this: BASE_DIR / 'subdir'.
BASE_DIR = Path(__file__).resolve().parent.parent
ANALYTICS_DIR = os.path.join(BASE_DIR, 'ClientAPI', 'analytics')

def create_excel_file(orders):
    # Проверяем, существует ли директория ANALYTICS_DIR, и если нет, то создаем ее
    if not os.path.exists(ANALYTICS_DIR):
        os.makedirs(ANALYTICS_DIR)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Увеличиваем размер столбцов
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15

    # Заголовки столбцов
    ws.append(['Номер заказа', 'Дата создания и время', 'Перечень', 'Итоговая сумма'])

    for order in orders:
        # Преобразуем дату и время к локальному времени и устанавливаем tzinfo в None
        local_datetime = localtime(order.datetimes).replace(tzinfo=None)

        # Добавляем данные в строки
        goods_list = order.goods.split('\n')  # Разбиваем строку на список товаров
        goods_column = '\n'.join(goods_list)  # Снова объединяем в столбец
        ws.append([order.order_id, local_datetime, goods_column, order.sum_price])

    # Добавляем формулу для суммирования колонки "Итоговая сумма"
    last_row = len(orders) + 2  # +2 для учета заголовка и формулы
    ws.cell(row=last_row, column=3, value=f'Итог:')
    ws.cell(row=last_row, column=4, value=f'=SUM(D2:D{last_row - 1})')

    # Выделяем только последнюю ячейку "Итоговая сумма" желтым цветом
    last_cell = ws.cell(row=last_row, column=4)
    last_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    last_cell_two = ws.cell(row=last_row, column=3)
    last_cell_two.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Создаем временный файл для хранения
    file_path = os.path.join(ANALYTICS_DIR, 'orders_report.xlsx')
    wb.save(file_path)

    return file_path

def download_excel_file(request):
    # Получаем все заказы
    orders = Order.objects.all()

    # Создаем Excel файл
    file_path = create_excel_file(orders)

    # Возвращаем JSON-ответ с путем к файлу
    response_data = {'file_path': file_path}
    return JsonResponse(response_data)