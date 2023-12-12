from typing import Text
import datetime
from django.db import models
from django.core.files.base import ContentFile
from openpyxl import Workbook
from django.utils.timezone import localtime
from openpyxl.styles import PatternFill
from io import BytesIO


offset = datetime.timezone(datetime.timedelta(hours=6))

class Category(models.Model):
    category_id: int = models.AutoField(primary_key=True)
    category_name: Text = models.TextField('category_name', max_length=255, unique=True, null=False,
                                           default='Новая категория')

    class Meta:
        db_table = 'Category'

    def __str__(self):
        return self.category_name


class Goods(models.Model):
    goods_id: int = models.AutoField(primary_key=True)
    title: Text = models.CharField('title', max_length=255, null=False, default='Без названия')
    description: Text = models.TextField('description', max_length=255,default='.')
    price_rub: int = models.FloatField('price_rub', null=False, default=0.00)
    category = models.ForeignKey('Category', on_delete=models.PROTECT)

    class Meta:
        db_table = 'Goods'

    def __str__(self):
        return f'{self.title}............. {self.price_rub}руб'


class Order(models.Model):
    order_id: int = models.AutoField(primary_key=True)
    goods = models.TextField('goods')
    datetimes: datetime = models.DateTimeField('datetimes', default=datetime.datetime.now(offset))
    sum_price: int = models.FloatField('sum_price', null=False, default=0.00)
    is_working: bool = models.BooleanField('is_working', default=False)  # This trigger if the value of the object -
    # is true,then the order will be displayed in the state 'At work'
    is_ready: bool = models.BooleanField('is_ready', default=False)  # This trigger if the value of the object -
    # is true, the order will be displayed in the 'Ready' state
    is_deleted: bool = models.BooleanField('is_deleted', default=False)  # This trigger if the value of the object -

    # is true, the order will not be displayed, and will be transferred to the archive
    class Meta:
        db_table = 'Order'

class Files(models.Model):
    file = models.FileField(upload_to='orders_files/', null=True, blank=True)

    class Meta:
        db_table = 'Files'

    def __str__(self):
        return f'File {self.id}'

    @staticmethod
    def create_excel_file(orders):
        # Создаем Excel-файл и сохраняем его в объект BytesIO
        content_file = BytesIO()
        wb = Workbook()
        ws = wb.active

        # Увеличиваем размер столбцов
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15

        # Заголовки столбцов
        ws.append(['Номер заказа', 'Дата создания и время', 'Перечень', 'Итоговая сумма'])

        for order in orders:
            # Создаем новую строку в каждой итерации цикла
            row = [
                order.order_id,
                localtime(order.datetimes).replace(tzinfo=None),
                '\n'.join(order.goods.split('\n')),
                order.sum_price
            ]
            ws.append(row)

        # Добавляем формулу для суммирования колонки "Итоговая сумма"
        last_row = len(orders) + 2  # +2 для учета заголовка и формулы
        ws.cell(row=last_row, column=3, value=f'Итог:')
        ws.cell(row=last_row, column=4, value=f'=SUM(D2:D{last_row - 1})')

        # Выделяем только последнюю ячейку "Итоговая сумма" желтым цветом
        last_cell = ws.cell(row=last_row, column=4)
        last_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        last_cell_two = ws.cell(row=last_row, column=3)
        last_cell_two.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Выводим содержимое orders и content_file для отладки
        print("Orders:", orders)
        print("Content File:", content_file.getvalue())

        # Сохраняем файл в объект BytesIO
        wb.save(content_file)

        # Сохраняем файл в модель Files
        obj = Files.objects.create()
        obj.file.save('orders_report.xlsx', ContentFile(content_file.getvalue()), save=True)

        return obj